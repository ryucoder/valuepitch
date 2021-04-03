from openpyxl import Workbook
from openpyxl.styles import Alignment


wb = Workbook()  
sheet = wb.active  
  

start_row = 5 

cell = sheet.cell(row=start_row, column=1)  
cell.value = 'Diary No'  
cell.alignment = Alignment(horizontal='center', vertical='center')  
  

sheet.merge_cells('B{}:C{}'.format(start_row, start_row))  
  

start_row += 1 

cell = sheet.cell(row=start_row, column=1)  
cell.value = 'Who Vs Who'  
cell.alignment = Alignment(horizontal='center', vertical='center')  

sheet.merge_cells('B{}:C{}'.format(start_row, start_row))  


start_row += 1 

cell = sheet.cell(row=start_row, column=1)  
cell.value = 'Case Details'  
cell.alignment = Alignment(horizontal='center', vertical='center')  

sheet.merge_cells('A{}:A{}'.format(start_row, start_row + 5))  


wb.save('merging.xlsx')  