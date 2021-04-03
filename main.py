import requests
import os
import csv

from pprint import pprint

import settings

from utils import HelperUtil

from bs4 import BeautifulSoup, Tag, NavigableString

from openpyxl import Workbook
from openpyxl.styles import Alignment


def initial_operations():

    # Condition 1
    HelperUtil.delete_folder_if_exists(settings.OUTPUT_FOLDER)

    # Condition 2
    HelperUtil.create_folder_if_doesnt_exists(settings.OUTPUT_FOLDER)

    # Any Other Conditions Go Below
    # pass


def download_responses():
    START = settings.START_YEAR
    STOP = settings.START_YEAR - settings.YEAR_RANGE
    RANGE = -1

    filenames = []

    for year_number in range(START, STOP, RANGE):
        for dairy_number in range(1, (settings.DAIRY_RANGE + 1)):

            response = requests.get(settings.CAPTCHA_URL)
            captcha_number = response.json()

            data = {
                "d_no": dairy_number,
                "d_yr": year_number,
                "ansCaptcha": captcha_number,
            }

            headers = {
                # 'Host': 'main.sci.gov.in',
                # 'Origin': 'https://main.sci.gov.in',
                "Referer": "https://main.sci.gov.in/case-status",
            }

            # Only Referer header is required to bypass CORS
            # You can verify using response.url
            response = requests.post(settings.TARGET_URL, data=data, headers=headers)

            # print()
            # print()
            # pprint(dir(response))
            # print(response.url)
            # print(response.status_code)
            # print(response.text) # str
            # print(response.content) # bytes
            # print()
            # print()

            # store the response to the output folder
            current_output_folder = os.path.join(
                settings.OUTPUT_FOLDER, str(year_number)
            )

            HelperUtil.create_folder_if_doesnt_exists(current_output_folder)

            current_filename = os.path.join(
                current_output_folder, "{}.html".format(dairy_number)
            )

            filenames.append(current_filename)

            with open(current_filename, "wb") as f:
                f.write(response.content)


def write_extracted_data_to_xlsx(extracted_data):
    print()
    print()
    print("write_extracted_data_to_xlsx")
    print()
    print()

    all_years = []
    for year in extracted_data.keys():
        all_years.append(year)

    # print()
    # print()
    # print("all_years")
    # print(all_years)
    # print()
    # print()

    wb = Workbook()  
    sheet = wb.active  
    
    start_row = 5 


    for year in range(len(all_years), 0, -1):
        latest_year = all_years[year - 1]
        print()
        print()
        print("year")
        print(latest_year)
        print()
        print()

        for item in extracted_data[latest_year]:


            cell = sheet.cell(row=start_row, column=1)  
            cell.value = 'Diary No'  
            cell.alignment = Alignment(horizontal='center', vertical='center')  
            

            sheet.merge_cells('B{}:C{}'.format(start_row, start_row))  
            

            start_row += 1 

            cell = sheet.cell(row=start_row, column=1)  
            cell.value = 'Who Vs Who'  
            cell.alignment = Alignment(horizontal='center', vertical='center')  

            sheet.merge_cells('B{}:C{}'.format(start_row, start_row))  



            start_row += 1 

            print()
            print()
            print("start_row")
            print(start_row)
            print()
            print()
            cell = sheet.cell(row=start_row, column=1)  
            cell.value = 'Case Details'  
            cell.alignment = Alignment(horizontal='center', vertical='center')  

            sheet.merge_cells('A{}:A{}'.format(start_row, start_row + 5))  

            print()
            print()
            print("item")
            print(item["diary_no"])
            print()
            print()
            item["diary_no"]
            item["who_vs_who"]

            cell = sheet.cell(row=start_row - 2, column=2)  
            cell.value = item["diary_no"]
            cell.alignment = Alignment(horizontal='center', vertical='center')  

            cell = sheet.cell(row=(start_row - 2) + 1, column=2)  
            cell.value = item["who_vs_who"]
            cell.alignment = Alignment(horizontal='center', vertical='center')  


            # item["Case Details"]
            start_row += 8


        # start_row += 1


    output_file = os.path.join(settings.OUTPUT_FOLDER, "output.xlsx")
    wb.save(output_file)


def extract_data_from_responses():
    """
    This is a very big method. It should be divided into smaller utils.
    As this is a testing task; I did not care to do so.
    """

    # read all the html files inside the output folder
    urls = {}

    for root, dirs, files in os.walk(settings.OUTPUT_FOLDER):
        if root != settings.OUTPUT_FOLDER:

            diary = int(root.split("/")[-1])
            # diary = str(root.split("/")[-1])

            urls[diary] = []

            for file_name in files:
                if file_name.split(".")[-1] == "html":
                    absolute_file_path = root + "/" + file_name
                    urls[diary].append(absolute_file_path)


    print()
    print()
    print("urls")
    pprint(urls)
    print()
    print()
    # loop over it and extract the data from each file and store in temp dict, put dict in list
    extracted_data = {}

    for diary_number in urls.keys():
        extracted_data[diary_number] = []

        for response_file in urls[diary_number]:

            one_case_file = {"diary_no": None, "who_vs_who": None}

            with open(response_file) as fp:
                soup = BeautifulSoup(fp, "html.parser")

                h5_tags = soup.find_all("h5")
                one_case_file["diary_no"] = h5_tags[0].string
                one_case_file["who_vs_who"] = h5_tags[1].string

                h4_tags = soup.find_all("h4")

                # Each case has different no of tabs
                for h4_tag in h4_tags:
                    one_case_file[h4_tag.find("a").string] = {}

                # print()
                # print()
                # print("one_case_file")
                # print(one_case_file)
                # print()
                # print()

                # Read Case Details
                one_case_file[h4_tags[0].find("a").string] = {}
                case_details = soup.find("div", {"id": "collapse1"})

                for table_row in case_details.find("table").find_all("tr"):

                    all_tds = table_row.find_all("td")

                    row_header = all_tds[0]
                    row_data = all_tds[1]

                    if len(row_data.contents) > 0:

                        contents = ""

                        for item in row_data.strings: 
                            if issubclass(NavigableString, type(item)):
                                contents += item.string + " "
                            
                        # print()
                        # print()
                        # print("contents")
                        # print(contents)
                        # print()
                        # print()

                    contents = contents.strip()

                    if contents == "":
                        contents = "None" 

                    if row_header.string != None:
                        one_case_file[h4_tags[0].find("a").string][row_header.string] = contents

            extracted_data[diary_number].append(one_case_file)

    print()
    print()
    print("extracted_data")
    # pprint(extracted_data)
    print()
    print()


    write_extracted_data_to_xlsx(extracted_data)

    return 

    for current_key in extracted_data.keys():

        output_file = os.path.join(
            settings.OUTPUT_FOLDER, current_key, "{}.csv".format(current_key)
        )

        for one_case_file in extracted_data[current_key]:

            row_list = []
            row_list.append(["Diary No", "Who Vs Who"])
            row_list.append([one_case_file["diary_no"], one_case_file["who_vs_who"]])
            row_list.append([])
            row_list.append(["Case Details"])

            all_keys = []
            all_values = []


            temp_keys = []
            for key in one_case_file["Case Details"].keys():
                temp_keys.append(key)

            print()
            print()
            print("temp_keys")
            print(temp_keys)
            print()
            print()
            print()

            temp_keys.sort()
            
            print()
            print()
            print("temp_keys")
            print(temp_keys)
            print()
            print()
            print()
            # for current_key, current_value in one_case_file["Case Details"].items():
            #     all_keys.append(current_key)
            #     all_values.append(current_value)

            for current_key in temp_keys:
                all_keys.append(current_key)
                all_values.append(one_case_file["Case Details"][current_key])

            row_list.append(all_keys)
            row_list.append(all_values)
            row_list.append([])
            row_list.append([])

            with open(output_file, "a", newline="") as file:
                writer = csv.writer(file)
                writer.writerows(row_list)


def main():
    # # Operation 1
    # initial_operations()
    # download_responses()

    # # Operation 2
    extract_data_from_responses()


if __name__ == "__main__":
    main()
