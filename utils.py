import os
import shutil

import settings

from bs4 import BeautifulSoup, Tag, NavigableString

from openpyxl import Workbook
from openpyxl.styles import Alignment


class HelperUtil:
    @staticmethod
    def delete_folder_if_exists(output_folder):
        if os.path.exists(output_folder):
            shutil.rmtree(output_folder)

    @staticmethod
    def create_folder_if_doesnt_exists(output_folder):
        if not os.path.exists(output_folder):
            os.mkdir(output_folder)

    @staticmethod
    def get_urls():
        # read all the html files inside the output folder
        urls = {}

        for root, dirs, files in os.walk(settings.OUTPUT_FOLDER):
            if root != settings.OUTPUT_FOLDER:

                diary = int(root.split("/")[-1])

                urls[diary] = []

                for file_name in files:
                    if file_name.split(".")[-1] == "html":
                        absolute_file_path = root + "/" + file_name
                        urls[diary].append(absolute_file_path)
        
        return urls

    @staticmethod
    def get_extracted_data(urls):

        extracted_data = {}

        for diary_number in urls.keys():
            extracted_data[diary_number] = []

            for response_file in urls[diary_number]:

                one_case_file = {"diary_no": None, "who_vs_who": None}

                with open(response_file) as fp:
                    soup = BeautifulSoup(fp, "html.parser")

                    h5_tags = soup.find_all("h5")
                    one_case_file["diary_no"] = h5_tags[0].string
                    one_case_file["who_vs_who"] = h5_tags[1].string

                    h4_tags = soup.find_all("h4")

                    # Each case has different no of tabs
                    for h4_tag in h4_tags:
                        one_case_file[h4_tag.find("a").string] = {}

                    # print()
                    # print()
                    # print("one_case_file")
                    # print(one_case_file)
                    # print()
                    # print()

                    # Read Case Details
                    one_case_file[h4_tags[0].find("a").string] = {}
                    case_details = soup.find("div", {"id": "collapse1"})

                    for table_row in case_details.find("table").find_all("tr"):

                        all_tds = table_row.find_all("td")

                        row_header = all_tds[0]
                        row_data = all_tds[1]

                        if len(row_data.contents) > 0:

                            contents = ""

                            for item in row_data.strings: 
                                if issubclass(NavigableString, type(item)):
                                    contents += item.string + " "
                                
                            # print()
                            # print()
                            # print("contents")
                            # print(contents)
                            # print()
                            # print()

                        contents = contents.strip()

                        if contents == "":
                            contents = "None" 

                        if row_header.string != None:
                            one_case_file[h4_tags[0].find("a").string][row_header.string] = contents

                extracted_data[diary_number].append(one_case_file)

        return extracted_data

    @staticmethod
    def write_extracted_data_to_xlsx(extracted_data):
        
        wb = Workbook()  
        sheet = wb.active  
        
        start_row = 5 

        all_years = []
        for year in extracted_data.keys():
            all_years.append(year)

        for year in range(len(all_years), 0, -1):
            latest_year = all_years[year - 1]

            for item in extracted_data[latest_year]:

                # First Header
                cell = sheet.cell(row=start_row, column=1)  
                cell.value = 'Diary No'  
                cell.alignment = Alignment(horizontal='center', vertical='center')  
                
                sheet.merge_cells('B{}:C{}'.format(start_row, start_row))  
                
                # Write to Cell
                cell = sheet.cell(row=start_row, column=2)  
                cell.value = item["diary_no"]
                cell.alignment = Alignment(horizontal='center', vertical='center')  


                start_row += 1 

                # Second Header
                cell = sheet.cell(row=start_row, column=1)  
                cell.value = 'Who Vs Who'  
                cell.alignment = Alignment(horizontal='center', vertical='center')  

                sheet.merge_cells('B{}:C{}'.format(start_row, start_row))  


                # Write to Cell
                cell = sheet.cell(row=start_row, column=2)  
                cell.value = item["who_vs_who"]
                cell.alignment = Alignment(horizontal='center', vertical='center')  


                start_row += 1 


                # Third Header
                cell = sheet.cell(row=start_row, column=1)  
                cell.value = 'Case Details'  
                cell.alignment = Alignment(horizontal='center', vertical='center')  

                column_nos = len(item['Case Details'])

                sheet.merge_cells('A{}:A{}'.format(start_row, start_row + column_nos - 1))  

                
                for key in item['Case Details'].keys():
                    value = item['Case Details'][key]

                    cell = sheet.cell(row=start_row, column=2)  
                    cell.value = key 
                    cell.alignment = Alignment(vertical='center')  

                    cell = sheet.cell(row=start_row, column=3)  
                    cell.value = item['Case Details'][key] 
                    cell.alignment = Alignment(vertical='center')  

                    # After each row in Case Details 
                    start_row += 1

                # After each entry
                start_row += 2

        output_file = os.path.join(settings.OUTPUT_FOLDER, "output.xlsx")
        wb.save(output_file)

